import tkinter as tk
import openpyxl as px
import subprocess as sp
import random, sys

#
#Excelファイルの読み込み
#
fileNum=int(input(
"""
<単語帳を選択してください>

1.応用情報技術者試験(AP)

2.未登録

3.未登録

"""))
if fileNum==1:
    fileName="AP.xlsx"
    filePath=r"C:\Users\taoka\Desktop\program\python\main\memorize_UI\AP.xlsx"
else:
    print("単語帳が存在しません")
    sys.exit()
wb=px.load_workbook(filePath)

#Unitの指定と読み込み
print("--------------------------------------------\n")
for index, name in enumerate(wb.sheetnames):
    print(str(index)+". "+name+"\n")
unitNum=int(input("""
Unitを選択してください：

--------------------------------------------
"""))
ws=wb.worksheets[unitNum]


#重複のない乱数生成（この後の作業で使用する）
def generateNumber(n1, n2, n3):  #どこから、どこまでで、何個生成するか
      if n2+1 < n3:
          n3=n2
      aa=0
      x=[]
      while aa < n3:
          n=random.randint(n1, n2)
          
          if n in x:
              continue
          else:
              x.append(n)
          
              aa+=1
      return x

#問題と解答リストの生成
question_list=[]
answer_list=[]
N=ws.max_row
k=generateNumber(2, N, N-1)
print(k)
print(N)

for i in k:
    question_list.append(ws.cell(row=i, column=1).value)
    answer_list.append(ws.cell(row=i, column=2).value)
#
#ボタン関数
#
key1=0
key2=0
def show_answer():
    global key1
    if key1+1==key2:
        txtBox3.delete(0, tk.END)
        txtBox3.insert(0, str(answer_list[key1]))
        key1+=1
        print("key1："+str(key1))
    
def add_list():
    ws_unknown=wb["未暗記リスト"]
    x=ws_unknown.max_row + 1
    list_num=k[key2-1]
    q=ws.cell(row=list_num, column=1).value
    a=ws.cell(row=list_num, column=2).value
    ws_unknown.cell(row=x, column=1).value=q
    ws_unknown.cell(row=x, column=2).value=a
    print(str(a)+q)
    wb.save(filePath)

    

def open_file():
    sp.Popen(["start", filePath], shell=True)

def next_question():
    global key2
    M=N-key2-1
    root.title("あと"+str(M)+"問")
    if key1==key2:
        txtBox1.delete(1.0, "end")
        txtBox2.delete(0, tk.END)
        txtBox3.delete(0, tk.END)
        txtBox1.insert(tk.END, str(question_list[key2]))
        key2+=1
        print("key2："+str(key2))
    


#コントロールエリア
#-----------------------------------------------
window_wid=600   #ウィンドウの横サイズ
window_hei=450   #ウィンドウの縦サイズ
question_wid=150  #問題欄の横サイズ
question_hei=50  #問題欄の縦サイズ
answer_wid=200    #回答欄の横サイズ
answer_hei=5     #回答欄の縦サイズ
button_wid=14    #ボタンの横サイズ
button_hei=2    #ボタンの縦サイズ
font1=("游ゴシック", 15, "bold")  #「問題」のフォント
font2=("游ゴシック", 12, "bold")  #「ボタン」のフォント
font3=("游ゴシック", 10, "bold")  #「回答」のフォント
#-----------------------------------------------

#
#ウィンドウUI作成
#
root=tk.Tk()
window_size=str(window_wid)+"x"+str(window_hei)
root.geometry(window_size)
root.resizable(width="False", height="False")
#オブジェクトの配置
label1=tk.Label(root, text="問題", font=font1)
label1.grid(row=0, column=0, padx=10, sticky=tk.S+tk.W)

txtBox1=tk.Text(root, height=8, wrap=tk.CHAR, font=font2)
txtBox1.grid(row=1, column=0, columnspan=2, padx=5, sticky=tk.W+tk.E)
txtBox1.insert("1.0", "「次へ」を押してください")

label2=tk.Label(root, text="回答：", font=font1)
label2.grid(row=2, column=0, padx=10, sticky=tk.W)

txtBox2=tk.Entry(root)
txtBox2.configure(width=90, font=font3)
txtBox2.grid(row=2, column=0, columnspan=2, padx=80, ipady=4, sticky=tk.W+tk.E)

label3=tk.Label(root, text="解答：", font=font1)
label3.grid(row=3, column=0, padx=10, sticky=tk.W)

txtBox3=tk.Entry(root)
txtBox3.configure(width=90, fg="red", font=font3)
txtBox3.grid(row=3, column=0, columnspan=2, padx=80, ipady=4, sticky=tk.W+tk.E)

answerButton=tk.Button(root, text="解答を表示", font=font2, command=show_answer)
answerButton.grid(row=4, column=0, padx=5, pady=5, sticky="NSEW")

addButton=tk.Button(root, text="未暗記リストに追加", font=font2, command=add_list)
addButton.grid(row=4, column=1, padx=5, pady=5, sticky="NSEW")

openButton=tk.Button(root, text="ファイルを開く", font=font2, command=open_file)
openButton.grid(row=5, column=0, padx=5, pady=5, sticky="NSEW")

nextButton=tk.Button(root, text="次へ", font=font2, command=next_question)
nextButton.grid(row=5, column=1, padx=5, pady=5, sticky="NSEW")

print("UIを作成した")

for i in range(6):
    root.rowconfigure(i, weight=1)
for i in range(2):
    root.columnconfigure(i, weight=1)

root.mainloop()